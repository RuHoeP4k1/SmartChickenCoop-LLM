"""
Export the 30 evaluation test cases as a formatted A4 Excel table.
Columns: # | Question | Ground Truth
"""
import sys
import os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "shared"))

from evaluation_data import TEST_CASES
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, Color
from openpyxl.utils import get_column_letter

OUT = os.path.join(os.path.dirname(__file__), "evaluation_questions.xlsx")

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Evaluation Questions"

# --- Page setup: A4 landscape ---
ws.page_setup.paperSize = ws.PAPERSIZE_A4
ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
ws.page_setup.fitToPage = True
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws.page_margins.left = 0.5
ws.page_margins.right = 0.5
ws.page_margins.top = 0.6
ws.page_margins.bottom = 0.6

# --- Styles ---
HEADER_FILL = PatternFill("solid", fgColor="2B4E72")   # dark blue
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
ROW_FONT    = Font(size=9, name="Calibri")
ALT_FILL    = PatternFill("solid", fgColor="EAF0F6")   # light blue-grey
WHITE_FILL  = PatternFill("solid", fgColor="FFFFFF")

thin = Side(style="thin", color="CCCCCC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

WRAP = Alignment(wrap_text=True, vertical="top")
CENTER_WRAP = Alignment(wrap_text=True, vertical="top", horizontal="center")

# --- Column widths (in Excel units, ~7 px each) ---
ws.column_dimensions["A"].width = 4    # #
ws.column_dimensions["B"].width = 30   # Question
ws.column_dimensions["C"].width = 75   # Ground Truth

# --- Header row ---
headers = ["#", "Question", "Ground Truth"]
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER_WRAP
    cell.border = BORDER
ws.row_dimensions[1].height = 22

# --- Data rows ---
for i, tc in enumerate(TEST_CASES, 1):
    row = i + 1
    fill = ALT_FILL if i % 2 == 0 else WHITE_FILL

    num_cell = ws.cell(row=row, column=1, value=i)
    num_cell.font = Font(size=9, bold=True, name="Calibri")
    num_cell.alignment = CENTER_WRAP
    num_cell.fill = fill
    num_cell.border = BORDER

    q_cell = ws.cell(row=row, column=2, value=tc["question"])
    q_cell.font = Font(size=9, bold=True, name="Calibri")
    q_cell.alignment = WRAP
    q_cell.fill = fill
    q_cell.border = BORDER

    gt_cell = ws.cell(row=row, column=3, value=tc["ground_truth"])
    gt_cell.font = ROW_FONT
    gt_cell.alignment = WRAP
    gt_cell.fill = fill
    gt_cell.border = BORDER

    # Estimate row height: ~14.5 pt per text line; ground_truth is the longest
    char_width_gt = 75 * 1.15  # approx chars per line at col width 75
    lines = max(1, len(tc["ground_truth"]) / char_width_gt * 1.5)
    ws.row_dimensions[row].height = max(14, lines * 14.5)

# Freeze header
ws.freeze_panes = "A2"

wb.save(OUT)
print(f"Saved: {OUT}")
print(f"Rows: {len(TEST_CASES)} questions")
